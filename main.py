
from openpyxl import load_workbook
wb2 = load_workbook('./data/raw_data.xlsx')
print(wb2.sheetnames)